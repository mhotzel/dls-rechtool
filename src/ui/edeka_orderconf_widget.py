from pathlib import Path
from typing import List, MutableMapping
import uuid
from PySide6.QtWidgets import (
    QGroupBox, QWidget, QFrame, QVBoxLayout,
    QPushButton, QLineEdit, QGridLayout,
    QFileDialog, QLabel
)
from PySide6.QtCore import (Signal, QThreadPool, QRunnable)

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from application.app_event import AppEvent, LogLevel
from application.event_dispatcher import EventDispatcher
from domain.event_factory import orderconfirmation_imported_event
from domain.order_confirmation import OrderConfirmation, OrderConfItem
from services.event_store.eventstore import EventStore, Event
from services.readmodels.base_data_store import DataStore, Document


class EdekaOrderConfirmationImportWidget(QGroupBox):

    statusSignal = Signal(AppEvent)

    def __init__(self, parent: QWidget, event_dispatcher: EventDispatcher, evtStore: EventStore, data_store: DataStore):
        super().__init__('EDEKA-Bestellbestätigungen importieren', parent)
        self.evtStore = evtStore
        self.data_store = data_store
        self.evt_dispatcher = event_dispatcher
        self.suppl_id = '1'
        self.threadPool = QThreadPool(self)
        self.statusSignal.connect(self.setStatusMessage)
        self.__build_ui()

    def setStatusMessage(self, msg: AppEvent) -> None:
        self.evt_dispatcher.send(msg)

    def __build_ui(self) -> None:
        """Baut die Oberfläche"""
        layout = QVBoxLayout(self)
        self.setLayout(layout)

        self.import_frame = QFrame(self)
        layout.addWidget(self.import_frame)
        layout.addStretch(1)

        self.import_frame.setLayout(QGridLayout(self.import_frame))
        self.btn_select_file = QPushButton(
            'Datei auswählen', parent=self.import_frame)

        txt = "ACHTUNG: Die Bestellbetätigungen werden "
        txt += f"'hart' der Lieferantennummer '{self.suppl_id}' zugeordnet!\n"
        txt += "EDEKA muss also in der Anwendung mit genau dieser Nummer angelegt sein"
        self.lbl_hinweis_liefnr = QLabel(txt, self.import_frame)
        self.lbl_hinweis_liefnr.setStyleSheet(
            '.QLabel {font-weight: bold; color: #CE0538}')

        self.btn_select_file.clicked.connect(lambda evt: self.select_file())
        self.txt_selected_file = QLineEdit(
            '', parent=self.import_frame, readOnly=True)

        self.btn_start_import = QPushButton(
            'Import starten', parent=self.import_frame)
        self.btn_start_import.clicked.connect(lambda evt: self.start_import())
        self.btn_start_import.setEnabled(False)

        self.import_frame.layout().addWidget(self.lbl_hinweis_liefnr, 0, 0)
        self.import_frame.layout().addWidget(self.btn_select_file, 1, 0)
        self.import_frame.layout().addWidget(self.txt_selected_file, 2, 0)
        self.import_frame.layout().addWidget(self.btn_start_import, 3, 0)

    def select_file(self):
        filename, ok = QFileDialog.getOpenFileName(
            self,
            caption='Bestellbestätigungsdatei auswählen',
            filter="Bestellungen-Datei (*.xlsx)"
        )

        if filename:
            self.txt_selected_file.setText(str(Path(filename)))
            self.btn_start_import.setEnabled(True)

    def start_import(self):
        """Startet den Import der Bestellbestätigungen"""

        docs=self.data_store.get_doc_list()

        thread = ImportWorker(
            self.run_import,
            filename=self.txt_selected_file.text(),
            evtStore=self.evtStore,
            docs=docs,
            statusSignal=self.statusSignal
        )

        self.threadPool.start(thread)

    def run_import(self, docs: List[Document], filename: str) -> None:
        wb: Workbook = load_workbook(
            filename,
            read_only=True, data_only=True)
        ws: Worksheet = wb.worksheets[0]
        item: OrderConfItem = None
        order_conf_map: MutableMapping[str, OrderConfirmation] = dict()

        try:
            for row in ws.iter_rows(3, min_col=1, max_col=11, values_only=True):
                if row[2] is None:
                    break
                order_confirmation_id = row[1].replace('.xlsx', '')
                if order_confirmation_id not in order_conf_map:
                    order_conf_map[order_confirmation_id] = OrderConfirmation(
                        suppl_id='1',
                        suppl_name='EDEKA',
                        order_confirm_id=order_confirmation_id,
                        order_date=row[8].date(),
                        positions=[]
                    )

                item = OrderConfItem(
                    idx=row[0],
                    seller_assigned_id=row[2],
                    global_id=row[3],
                    name=row[4],
                    quantity=row[5],
                    unitcode=row[6],
                    packaging_quantity=row[7],
                    price=row[9],
                    total_line_amount=row[10]
                )
                order_conf_map[order_confirmation_id].positions.append(item)
                self.statusSignal.emit(
                    AppEvent(
                        evt_lvl=LogLevel.INFO,
                        evt_type='status-message',
                        evt_data=f"Bestellbestätigung '{item.idx}' wurde eingelesen"
                    )
                )

            for key, orderconf in order_conf_map.items():
                evt = orderconfirmation_imported_event(orderconf)

                for doc in docs:
                    if doc.subject == evt.subject:
                        raise ValueError(f"Das Dokument '{evt.subject}' wurde bereits eingelesen")

                self.evtStore.add_event(evt=evt, expected_version=None)
                self.statusSignal.emit(
                    AppEvent(
                        evt_lvl=LogLevel.INFO,
                        evt_type='status-message',
                        evt_data="Bestellbestätigung '{item.idx}' wurden importiert"
                    )
                )

            self.statusSignal.emit(
                AppEvent(
                    evt_lvl=LogLevel.INFO,
                    evt_type='status-message',
                    evt_data='Alle Bestellbestätigungen wurden importiert'
                )
            )
        except Exception as e:
            print(repr(e))
            self.statusSignal.emit(
                AppEvent(
                    evt_lvl=LogLevel.CRITICAL,
                    evt_type='status-message',
                    evt_data=f"Import war fehlerhaft bei '{item}': {e}"
                )
            )


class ImportWorker(QRunnable):

    def __init__(self, fn, filename: str, evtStore: EventStore, docs: List[Document], statusSignal: Signal):
        super().__init__()
        self.fn = fn
        self.evtStore = evtStore
        self.docs = docs
        self.filename = filename
        self.statusSignal = statusSignal

    def run(self) -> None:
        self.fn(self.docs, self.filename)
