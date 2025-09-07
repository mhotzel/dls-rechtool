
from datetime import datetime
from pathlib import Path
from typing import List
from PySide6.QtWidgets import (
    QGroupBox, QWidget, QVBoxLayout, QFrame, QPushButton, QLineEdit, QFileDialog
)
from PySide6.QtCore import (Signal, QThreadPool, QRunnable)

from application.app_event import AppEvent, LogLevel
from application.event_dispatcher import EventDispatcher
from services.readmodels.base_data_store import DataStore, Product
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

class ExportWorker(QRunnable):

    def __init__(self, fn, folder_with_filename: str, data_repo: DataStore, statusSignal: Signal):
        super().__init__()
        self.fn = fn
        self.data_repo: DataStore = data_repo
        self.filename = folder_with_filename
        self.statusSignal = statusSignal

    def run(self) -> None:
        self.fn(self.filename, self.data_repo)

class ProdListExportWidget(QGroupBox):
    """Bietet die Möglichkeit, die Produktliste zu exportieren."""

    statusSignal = Signal(AppEvent)

    def __init__(self, parent: QWidget, event_dispatcher: EventDispatcher, data_repo: DataStore):
        super().__init__('Einkaufspreisliste exportieren', parent, )
        self.evt_dispatcher = event_dispatcher
        self.data_repo: DataStore = data_repo

        self.threadPool = QThreadPool(self)
        self.statusSignal.connect(self.setStatusMessage)
        self.__build_ui()

    def __build_ui(self) -> None:
        """Baut die Oberfläche"""
        layout = QVBoxLayout()
        self.setLayout(layout)

        self.export_frame = QFrame(self)
        self.export_frame.setLayout(QVBoxLayout(self.export_frame))
        self.btn_select_folder = QPushButton(
            'Pfad für Exportdatei auswählen', parent=self.export_frame)
        self.btn_select_folder.clicked.connect(lambda evt: self.select_export_folder())
        self.txt_selected_folder = QLineEdit(
            '', parent=self.export_frame, readOnly=True)

        self.btn_start_export = QPushButton(
            'Import starten', parent=self.export_frame)
        self.btn_start_export.clicked.connect(lambda evt: self.start_export())
        self.btn_start_export.setEnabled(False)
        self.export_frame.layout().addWidget(self.btn_select_folder)
        self.export_frame.layout().addWidget(self.txt_selected_folder)
        self.export_frame.layout().addWidget(self.btn_start_export)

        layout.addWidget(self.export_frame)
        layout.addStretch(1)

    def select_export_folder(self):
        home = Path.home()
        folder = QFileDialog.getExistingDirectory(
            self,
            caption='Pfad für Export auswählen',
            dir=str(home.absolute())
        )

        if folder:
            datum = datetime.now().strftime('%Y%m%d-%H%M%S')
            path = Path(folder) / f"DLS-Artikel-EK-Liste_{datum}.xlsx"
            self.txt_selected_folder.setText(str(path))
            self.btn_start_export.setEnabled(True)

    def setStatusMessage(self, msg: AppEvent) -> None:
        self.evt_dispatcher.send(msg)

    def run_export(self, filename: str, data_store: DataStore) -> None:

        result: List[Product] = self.data_repo.get_product_list()

        wb = Workbook()
        ws: Worksheet = wb.active
        ws.title = "DLS-EK-Preise"

        ws.append(('Lief-Nr', 'Lief-Name', 'Typ', 'ID Herk.', 'Datum Eingang', 'Art-Nr', 'GTIN', 'Art-Name', 'EK Netto'))

        for row in result:
            ws.append((row.suppl_id, row.suppl_name, row.issue_type, row.issue_id, row.issue_date, row.seller_assigned_id, row.global_id, row.name, row.price))

        ws.freeze_panes = "A2"

        cell: Cell = None
        for cell in ws["I"]: 
            if cell.row == 1:  # Kopfzeile überspringen
                continue
            cell.number_format = '#,##0.000 €'  # deutsches Format

        for cell in ws["E"]: 
            if cell.row == 1:  # Kopfzeile überspringen
                continue
            cell.number_format = 'DD.MM.YYYY'  # deutsches Format

        for cell in ws[1]:   # ws[1] greift auf die gesamte erste Zeile zu
            cell.font = Font(bold=True)
            cell.number_format = "@"

        wb.save(filename=filename)
        
        # Nachricht, wenn fertig
        self.statusSignal.emit(
            AppEvent(
                evt_lvl=LogLevel.INFO,
                evt_type='status-message',
                evt_data=f"Ausgabedatei '{filename}' wurde erstellt"
            )
        )

    def start_export(self):
        """Startet den Export"""

        path = Path(self.txt_selected_folder.text())

        thread = ExportWorker(
            self.run_export,
            folder_with_filename=path,
            data_repo=self.data_repo,
            statusSignal=self.statusSignal
        )

        self.threadPool.start(thread)