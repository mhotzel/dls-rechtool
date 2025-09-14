
import datetime
from typing import List, Union
from openpyxl import load_workbook, Workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from domain.event_factory import GenericDocPosition, GenericDocument, GenericInvoice, GenericInvoicePosition, GenericOrder, GenericOrderPosition
from services.readmodels.base_data_store import DataStore


class LoadGenericDocumentCmd:
    """Command zum Laden einer Exceldatei mit Positionen"""

    def __init__(self, data_store: DataStore, filename: str):
        self.filename = filename
        self.data_store = data_store
        self.suppliers = []

    def load(self) -> GenericDocument:
        """Lädt die Datei und liefert ein Dokument zurück"""

        try:
            wb: Workbook = load_workbook(
                self.filename, data_only=True, read_only=True)
            self.suppliers = self.data_store.get_suppliers_list()
            self._check_format(wb)

            ws: Worksheet = wb["Eingabe"]
            return self.load_document(ws)
        except Exception as e:
            raise e
        finally:
            wb.close()

    def load_document(self, ws: Worksheet) -> GenericDocument:
        """Lädt Rechnungen"""

        positions: List[GenericDocument] = []
        suppl_id = ws.cell(2, 1).value

        suppl_dict = {sup.suppl_id: sup.suppl_name for sup in self.suppliers}

        doc = GenericDocument(
            doctype=ws.cell(2, 2).value,
            suppl_id=ws.cell(2, 1).value,
            suppl_name=suppl_dict[suppl_id],
            doc_id=ws.cell(2, 3).value,
            doc_date=ws.cell(2, 4).value,
            positions=None
        )

        pos: GenericDocPosition = None
        for counter, row in enumerate(ws.iter_rows(min_row=5, min_col=1, max_col=5, values_only=True)):
            if row[0] is None or row[0] == '':
                break

            pos = GenericDocPosition(idx=counter, line_id=str(
                row[0]), sellerAssignedId=row[1], globalId=row[3], name=row[2], price=row[4])
            positions.append(pos)

        doc.positions = positions
        return doc

    def _check_format(self, wb: Workbook) -> None:
        if "Eingabe" not in wb.sheetnames:
            raise ValueError(
                "Kein gültiges Dokument: Es ist kein Arbeitsblatt mit dem Namen 'Eingabe' vorhanden")

        ws: Worksheet = wb['Eingabe']
        if ws.cell(1, 1).value != 'Lief-Nr':
            raise ValueError("Kein gültiges Dokument")

        suppl_ids = [sup.suppl_id for sup in self.suppliers]
        suppl_id = ws.cell(2, 1).value
        if suppl_id not in suppl_ids:
            raise ValueError(f"Die Lieferanten-Nr. '{suppl_id}' in der Eingabedatei ist nicht bekannt. Bitte prüfen und ggf. nachpflegen")

        if ws.cell(1, 2).value != 'Dokumenttyp':
            raise ValueError("Kein gültiges Dokument")

        if ws.cell(1, 3).value != 'Dokument-ID':
            raise ValueError("Kein gültiges Dokument")

        if ws.cell(1, 4).value != 'Dokument-Datum':
            raise ValueError("Kein gültiges Dokument")

        if ws.cell(4, 1).value != 'Pos-Nr':
            raise ValueError("Kein gültiges Dokument")

        if ws.cell(4, 2).value != 'Artikel-Nr. Lieferant':
            raise ValueError("Kein gültiges Dokument")

        if ws.cell(4, 3).value != 'Artikelbez.':
            raise ValueError("Kein gültiges Dokument")

        if ws.cell(4, 4).value != 'GTIN / EAN':
            raise ValueError("Kein gültiges Dokument")

        if ws.cell(4, 5).value != 'Einzelpreis':
            raise ValueError("Kein gültiges Dokument")

        if ws.cell(2, 2).value not in ('Rechnung', 'Bestellung'):
            raise ValueError(
                "Die Zelle 'Dokumenttyp' muss 'Bestellung' oder 'Rechnung' enthalten")

        daten = ws.cell(2, 4).value
        if type(daten) != datetime.datetime:
            raise ValueError(
                "Die Zelle 'Dokument-Datum' enthält kein gültiges Datum")
